# analytics.py
"""
Main script for Assignment #2:
- Executes SQL queries (with JOINs)
- Creates 6 chart types and saves to /charts/
- Creates a Plotly time-slider animation (not saved but will also save HTML)
- Exports multiple sheets to Excel with formatting (openpyxl)
- Prints console reports required by assignment
"""
import os
import pandas as pd
import matplotlib.pyplot as plt
import plotly.express as px
from sqlalchemy import create_engine
import matplotlib.ticker as ticker
from config import DB_URL, CHARTS_FOLDER, EXPORTS_FOLDER
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

# Ensure output folders exist
os.makedirs(CHARTS_FOLDER, exist_ok=True)
os.makedirs(EXPORTS_FOLDER, exist_ok=True)

engine = create_engine(DB_URL)

def run_query(sql):
    df = pd.read_sql_query(sql, engine)
    print(f"Query returned {len(df):,} rows.")
    return df

# --- Подключение к БД ---
engine = create_engine("sqlite:///f1.db")

# 1) Pie chart: distribution of race wins by constructor (2010-2020)
pie_sql = """
SELECT c.name AS constructor, COUNT(*) AS wins
FROM results r
JOIN constructors c ON r.constructorId = c.constructorId
JOIN races ra ON r.raceId = ra.raceId
WHERE r.position = 1
  AND ra.year BETWEEN 2010 AND 2020
GROUP BY c.name
ORDER BY wins DESC;
"""
pie_df = run_query(pie_sql)

if not pie_df.empty:
    # --- Считаем доли ---
    total = pie_df["wins"].sum()
    pie_df["share"] = pie_df["wins"] / total

    # --- Делим на большие и маленькие сегменты ---
    large = pie_df[pie_df["share"] >= 0.01]   # ≥1%
    small = pie_df[pie_df["share"] < 0.01]    # <1%

    # --- Добавляем категорию "Other (<1%)" ---
    if not small.empty:
        other_row = pd.DataFrame({
            "constructor": ["Other (<1%)"],
            "wins": [small["wins"].sum()],
            "share": [small["share"].sum()]
        })
        pie_plot_df = pd.concat([large, other_row], ignore_index=True)
    else:
        pie_plot_df = large

    # --- Строим круговую диаграмму ---
    fig, ax = plt.subplots(figsize=(8, 8))
    ax.pie(
        pie_plot_df["wins"],
        labels=pie_plot_df["constructor"],
        autopct="%1.1f%%",
        startangle=140
    )
    ax.set_title("Share of Race Wins by Constructor (2010-2020)")

    # --- Если есть маленькие сегменты — таблица снизу ---
    if not small.empty:
        table_data = small[["constructor", "wins", "share"]].copy()
        table_data["share"] = (table_data["share"] * 100).round(2).astype(str) + "%"
        plt.table(
            cellText=table_data.values,
            colLabels=table_data.columns,
            loc="bottom",
            bbox=[0.0, -0.35, 1, 0.25]
        )
        plt.subplots_adjust(bottom=0.35)

    # --- Сохраняем картинку ---
    path = os.path.join(CHARTS_FOLDER, "pie_wins_by_constructor_2010_2020.png")
    fig.savefig(path, bbox_inches='tight')
    plt.close(fig)
    print(f"Saved pie chart: {path} — shows share of race wins by constructor 2010-2020 (with <1% grouped as 'Other').")

# 2) Bar chart: Top 10 drivers by total points (all years)
bar_sql = """
SELECT d.driverRef AS driver, SUM(r.points) AS total_points
FROM results r
JOIN drivers d ON r.driverId = d.driverId
JOIN races ra ON r.raceId = ra.raceId
GROUP BY d.driverRef
ORDER BY total_points DESC
LIMIT 10;
"""
bar_df = run_query(bar_sql)
if not bar_df.empty:
    ax = bar_df.plot.bar(x='driver', y='total_points', legend=False, title='Top 10 Drivers by Total Points (all years)', figsize=(10,6))
    ax.set_ylabel('Total Points')
    path = os.path.join(CHARTS_FOLDER, "bar_top10_drivers_points.png")
    fig = ax.get_figure()
    fig.savefig(path, bbox_inches='tight')
    plt.close(fig)
    print(f"Saved bar chart: {path} — shows top 10 drivers by cumulative points.")

# 3) Horizontal bar: Average grid position by constructor in 2015
hbar_sql = """
SELECT c.name AS constructor, AVG(r.grid) AS avg_grid
FROM results r
JOIN constructors c ON r.constructorId = c.constructorId
JOIN races ra ON r.raceId = ra.raceId
WHERE ra.year = 2015
  AND r.grid > 0
GROUP BY c.name
ORDER BY avg_grid ASC
LIMIT 15;
"""
hbar_df = run_query(hbar_sql)
if not hbar_df.empty:
    ax = hbar_df.plot.barh(x='constructor', y='avg_grid', legend=False, title='Average Grid Position by Constructor (2015)', figsize=(10,6))
    ax.set_xlabel('Average Grid Position')
    path = os.path.join(CHARTS_FOLDER, "hbar_avg_grid_constructor_2015.png")
    fig = ax.get_figure()
    fig.savefig(path, bbox_inches='tight')
    plt.close(fig)
    print(f"Saved horizontal bar chart: {path} — shows average qualifying grid position per constructor in 2015.")
#4
champ_sql = """
SELECT ra.year, d.driverRef AS driver, SUM(r.points) AS total_points
FROM results r
JOIN drivers d ON r.driverId = d.driverId
JOIN races ra ON r.raceId = ra.raceId
WHERE ra.year BETWEEN 2000 AND 2020
GROUP BY ra.year, d.driverRef
ORDER BY ra.year, total_points DESC;
"""
champ_df = run_query(champ_sql)

if not champ_df.empty:
    # pick champion per year (driver with max total_points)
    champ_per_year = (
        champ_df.sort_values(['year','total_points'], ascending=[True, False])
        .groupby('year')
        .first()
        .reset_index()
    )

    # преобразуем year в int
    champ_per_year["year"] = champ_per_year["year"].astype(int)

    ax = champ_per_year.plot.line(
        x='year',
        y='total_points',
        marker='o',
        title='Season Champion Points by Year (2000-2020)',
        figsize=(10,6)
    )
    ax.set_ylabel('Points')

    # 👉 делаем ось X только с целыми числами (убираем .0)
    ax.xaxis.set_major_locator(ticker.MaxNLocator(integer=True))

    path = os.path.join(CHARTS_FOLDER, "line_champion_points_2000_2020.png")
    fig = ax.get_figure()
    fig.savefig(path, bbox_inches='tight')
    plt.close(fig)
    print(f"Saved line chart: {path} — shows season champion total points per year (2000-2020).")


# 5) Histogram: distribution of lap times (milliseconds) for 2015-2019
hist_sql = """
SELECT lt.milliseconds
FROM lap_times lt
JOIN races ra ON lt.raceId = ra.raceId
JOIN drivers d ON lt.driverId = d.driverId
WHERE ra.year BETWEEN 2015 AND 2019
  AND lt.milliseconds IS NOT NULL;
"""
hist_df = run_query(hist_sql)
if not hist_df.empty:
    ax = hist_df['milliseconds'].plot.hist(bins=30, title='Distribution of Lap Times (ms) — 2015-2019', figsize=(10,6))
    ax.set_xlabel('Lap time (ms)')
    path = os.path.join(CHARTS_FOLDER, "hist_lap_times_2015_2019.png")
    fig = ax.get_figure()
    fig.savefig(path, bbox_inches='tight')
    plt.close(fig)
    print(f"Saved histogram: {path} — shows distribution of lap times (2015-2019).")

# 6) Scatter: Grid (qualifying position) vs Finishing position (sample)
scatter_sql = """
SELECT r.grid, r.position
FROM results r
JOIN drivers d ON r.driverId = d.driverId
JOIN races ra ON r.raceId = ra.raceId
WHERE r.grid > 0 AND r.position > 0
LIMIT 10000;
"""
scatter_df = run_query(scatter_sql)
if not scatter_df.empty:
    ax = scatter_df.plot.scatter(x='grid', y='position', title='Grid Position vs Finishing Position (sample)', figsize=(8,6))
    ax.set_xlabel('Grid (qualifying position)')
    ax.set_ylabel('Finishing position')
    path = os.path.join(CHARTS_FOLDER, "scatter_grid_vs_position.png")
    fig = ax.get_figure()
    fig.savefig(path, bbox_inches='tight')
    plt.close(fig)
    print(f"Saved scatter plot: {path} — shows relationship between starting grid and finishing position (sample).")

# 2) Plotly Time Slider (animation) — wins per constructor per year (2000-2020)
anim_sql = """
SELECT ra.year, c.name AS constructor, COUNT(*) AS wins
FROM results r
JOIN constructors c ON r.constructorId = c.constructorId
JOIN races ra ON r.raceId = ra.raceId
WHERE r.position = 1 AND ra.year BETWEEN 2000 AND 2020
GROUP BY ra.year, c.name
ORDER BY ra.year, wins DESC;
"""
anim_df = run_query(anim_sql)
if not anim_df.empty:
    # For years where a constructor has 0 wins, it won't appear; Plotly animation works fine with what we have.
    fig = px.bar(anim_df, x='constructor', y='wins', color='constructor',
                 animation_frame='year', range_y=[0, anim_df['wins'].max()+1],
                 title='Wins per Constructor by Year (2000-2020)')
    html_path = os.path.join(CHARTS_FOLDER, "plotly_wins_by_constructor_time_slider.html")
    fig.write_html(html_path)
    print(f"Created Plotly animation (time slider). Open {html_path} in a browser to interact during defense.")

# 3) Export to Excel with formatting
def export_to_excel(dfs_dict, filename):
    outpath = os.path.join(EXPORTS_FOLDER, filename)
    with pd.ExcelWriter(outpath, engine='openpyxl') as writer:
        for sheet_name, df in dfs_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    # openpyxl formatting
    wb = load_workbook(outpath)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # freeze header row
        ws.freeze_panes = "A2"
        # auto filter
        ws.auto_filter.ref = ws.dimensions
        # apply a 3-color scale to numeric columns
        # find numeric columns by checking first data row types
        max_col = ws.max_column
        max_row = ws.max_row
        for col_idx in range(1, max_col+1):
            col_letter = get_column_letter(col_idx)
            # check if the column has numeric values in the second row
            cell = ws[f"{col_letter}2"]
            try:
                val = cell.value
                is_numeric = isinstance(val, (int, float))
            except:
                is_numeric = False
            if is_numeric:
                # apply ColorScaleRule to C2:C{max_row} etc.
                rng = f"{col_letter}2:{col_letter}{max_row}"
                rule = ColorScaleRule(start_type="min", start_color="FFAA0000",
                                      mid_type="percentile", mid_value=50, mid_color="FFFFFF00",
                                      end_type="max", end_color="FF00AA00")
                ws.conditional_formatting.add(rng, rule)
    wb.save(outpath)
    # summary
    total_rows = sum([len(df) for df in dfs_dict.values()])
    print(f"Created file {filename}, {len(dfs_dict)} sheets, {total_rows} rows (saved to {outpath}).")
    return outpath

dfs_to_export = {
    "wins_by_constructor_2010_2020": pie_df,
    "top10_drivers_points": bar_df,
    "avg_grid_constructor_2015": hbar_df,
    "champion_points_2000_2020": champ_per_year if 'champ_per_year' in locals() else pd.DataFrame()
}

export_to_excel(dfs_to_export, "f1_report.xlsx")
